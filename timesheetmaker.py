from copy import copy
from openpyxl import *

# opening the source excel file
filename ="D://lnbtimesheetassembler/default.xlsx"
wb1 = load_workbook(filename)
ws1 = wb1.worksheets[0]
  
# opening the destination excel file 



wb2 = Workbook()
ws2 = wb2.active
for row in ws1.rows:
	for cell in row:
		new_cell = ws2.cell(row=cell.row, column=cell.column, value=cell.value)

		if cell.has_style:
			new_cell.font = copy(cell.font)
			new_cell.border = copy(cell.border)
			new_cell.fill = copy(cell.fill)
			new_cell.number_format = copy(cell.number_format)
			new_cell.protection = copy(cell.protection)
			new_cell.alignment = copy(cell.alignment)
wb2.save("Created Timesheet.xlsx")

  
# calculate total number of rows and 
# columns in source excel file
mr = ws1.max_row
mc = ws1.max_column
  




